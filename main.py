"""
main.py — sve u jednom: konfiguracija, logika i ulazna točka
============================================================

Pokretanje:
    python main.py                 # GUI
    python main.py gui             # GUI
    python main.py cli slika.jpg   # terminalna analiza
    python main.py train vgg16     # treniranje: cnn / vgg16 / oba
    python main.py eval            # usporedba modela

Datoteke projekta:
    main.py        konfiguracija + podaci + modeli + Grad-CAM + inferencija +
                   izvještaji + CLI ulaz   (zajednička jezgra za GUI i CLI)
    treniranje.py  treniranje modela + terminalni prikaz
    evaluacija.py  metrike, matrica konfuzije, grafovi, usporedba modela
    gui.py         grafičko sučelje

Napomena: gui.py / treniranje.py / evaluacija.py pristupaju ovome preko
`import main as cfg` (konstante) i `import main as core` (funkcije).
"""

import os
import io
import argparse
import datetime

import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
from PIL import Image
from sklearn.model_selection import train_test_split
from torch.utils.data import Dataset, DataLoader
from torchvision import transforms, models
from torchvision.models import VGG16_Weights

import matplotlib
matplotlib.use("Agg")


# ================================================================
# PUTANJE
# ================================================================
DIR_PODACI = "Data_split"
DIR_MODELI = "modeli"
DIR_REZULTATI = "rezultati"
DIR_REZULTATI_TRENIRANJA = "Rezultati treniranja"

PUTANJA_MODEL = {
    "cnn": os.path.join(DIR_MODELI, "cnn_best.pth"),
    "vgg16": os.path.join(DIR_MODELI, "vgg16_best.pth"),
}

# Izvor podataka: "imagefolder" (Data/train/NORMAL...) ili "rsna" (slike + CSV s Target)
IZVOR_PODATAKA = "rsna"
TEST_UDIO = 0.15                # RSNA: udio (grouped) koji ide u test

# ================================================================
# KLASE
# ================================================================
KLASE = ["NORMAL", "PNEUMONIA"]
KLASE_PRIKAZ = ["NORMALNO", "UPALA PLUĆA"]
POZITIVNA_KLASA = 1
NAZIV_POZITIVNE = "PNEUMONIA"

# ================================================================
# SLIKE / PRETPROCESIRANJE
# ================================================================
IMG_SIZE = 224
MEAN = [0.485, 0.456, 0.406]
STD = [0.229, 0.224, 0.225]
PODRZANI_FORMATI = (".png", ".jpg", ".jpeg", ".bmp", ".tiff")

# ================================================================
# TRENING
# ================================================================
BATCH = 32
NUM_WORKERS = 4
SJEME = 42
EPOCHS = 20
LR_CNN = 1e-3
LR_VGG = 1e-4
WEIGHT_DECAY = 1e-4
VAL_UDIO = 0.10
ZAMRZNI_VGG_BAZU = True
KORISTI_AMP = True
EARLY_STOP_STRPLJENJE = 5

# ---- VGG dvofazno fine-tuniranje (faza 1: glava; faza 2: odmrznuta baza) ----
VGG_FINE_TUNE = True            # True -> nakon faze 1 odmrzni dio baze i fino tuniraj
VGG_ODMRZNI_NA = 6             # epoha na kojoj kreće faza 2
VGG_ODMRZNI_BLOKOVA = 1        # koliko zadnjih conv blokova odmrznuti (1-2 sigurnije)
FINE_TUNE_LR = 1e-5           # nizak LR u fazi 2

# ---- Top-N modela + ručni izbor u konzoli ----
TOP_N_MODELA = 5
IZBOR_MODELA = True            # True -> nakon epoha pitaj koji od top-N spremiti kao best
SPREMI_SVE_TOP = False         # True -> spremi i svih top-N na disk (VGG je velik ~0.5GB/komad!)

# ================================================================
# SEGMENTACIJA PLUĆA (opcionalno pretprocesiranje)
# ================================================================
# True  -> prije svega se slika izreže na područje pluća (bbox iz maske).
#          Treba: pip install torchxrayvision    (skine težine segmentatora)
#          PRVO provjeri maske:  python segmentacija.py provjera Data/test/PNEUMONIA
# False -> standardno (CenterCrop). Ako segmentacija zakaže, automatski padne na False.
SEGMENTACIJA = False

# ================================================================
# ODLUKA / METRIKE
# ================================================================
PRAG_ODLUKE = 0.50
# Prag se nakon treniranja bira automatski (Youdenova točka: max recall+spec−1).
# RECALL_POD: opcionalni donji prag osjetljivosti (medicinski kontekst); None = bez poda.
RECALL_POD = 0.85
# Balans klasa pri treniranju: "sampler" (WeightedRandomSampler) | "loss" (težine u lossu) | "ne"
BALANS_KLASA = "sampler"

# ================================================================
# GRAD-CAM
# ================================================================
GRADCAM_ALPHA = 0.45
GRADCAM_KLASA = "predvidjena"   # "predvidjena" ili "pneumonija"

# ================================================================
# UREĐAJ
# ================================================================
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
if DEVICE == "cuda":
    torch.backends.cudnn.benchmark = True   # brže za fiksnu veličinu ulaza


# ================================================================
# PODACI, MODELI, GRAD-CAM, INFERENCIJA, IZVJEŠTAJI
# ================================================================
# ================================================================
# DATASET I TRANSFORMACIJE
# ================================================================
class XrayDataset(Dataset):
    def __init__(self, uzorci, transform):
        self.uzorci = uzorci
        self.transform = transform

    def __len__(self):
        return len(self.uzorci)

    def __getitem__(self, i):
        putanja, oznaka = self.uzorci[i]
        slika = Image.open(putanja).convert("L")
        import segmentacija
        slika = segmentacija.mozda_izrezi(slika, putanja)
        return self.transform(slika), oznaka


def napravi_transforme(img_size=None, augmentacija=False):
    """
    Anti-"Clever Hans" pretprocesiranje.

    Model je prije hvatao artefakte IZVAN pluća (rubovi kadra, "R" marker,
    pozicija) umjesto plućnog tkiva. Zato:
      • resize na malo veće pa CenterCrop -> rubovi i marker ispadaju iz ulaza
        (deterministički; primjenjuje se i na validaciju i na inferenciju),
      • u treningu RandomResizedCrop + ColorJitter + RandomErasing -> mreža se
        ne može osloniti na fiksnu poziciju/svjetlinu/kutove.
    """
    img_size = img_size or IMG_SIZE
    veci = int(round(img_size * 1.15))

    pre = [transforms.Grayscale(num_output_channels=3)]
    if augmentacija:
        geo = [
            transforms.Resize((veci, veci)),
            transforms.RandomResizedCrop(img_size, scale=(0.8, 1.0), ratio=(0.9, 1.1)),
            transforms.RandomHorizontalFlip(p=0.5),
            transforms.RandomRotation(degrees=10),
            transforms.ColorJitter(brightness=0.2, contrast=0.2),
        ]
    else:                                   # validacija + inferencija
        geo = [
            transforms.Resize((veci, veci)),
            transforms.CenterCrop(img_size),   # makni rubove + "R" marker
        ]

    zavrsne = [transforms.ToTensor(), transforms.Normalize(MEAN, STD)]
    if augmentacija:                        # cutout radi na tenzoru -> nakon ToTensor
        zavrsne.append(transforms.RandomErasing(p=0.3, scale=(0.02, 0.08)))

    return transforms.Compose(pre + geo + zavrsne)


def transform_inferencija(img_size=None):
    return napravi_transforme(img_size, augmentacija=False)


def geometrija_prikaza(pil, img_size=None):
    """Slika u ISTOJ geometriji koju model vidi (Resize veci + CenterCrop), kao RGB.
    Koristi se kao baza za Grad-CAM overlay -> heatmap se poravna s onim sto je
    model stvarno gledao (rješava kvadrat-vs-nekvadrat pomak)."""
    img_size = img_size or IMG_SIZE
    veci = int(round(img_size * 1.15))
    g = transforms.Compose([
        transforms.Grayscale(num_output_channels=3),
        transforms.Resize((veci, veci)),
        transforms.CenterCrop(img_size),
    ])
    return g(pil)


def _rijesi_korijen(dir_podaci):
    if os.path.isdir(os.path.join(dir_podaci, "train")):
        return dir_podaci
    ugnijezden = os.path.join(dir_podaci, "chest_xray")
    if os.path.isdir(os.path.join(ugnijezden, "train")):
        return ugnijezden
    raise FileNotFoundError(
        f"Ne nalazim 'train' podmapu u '{dir_podaci}'. Provjeri da postoji "
        f"{dir_podaci}/train/NORMAL i {dir_podaci}/train/PNEUMONIA."
    )


def _skupi(mapa):
    uzorci = []
    for oznaka, klasa in enumerate(KLASE):
        dir_klase = os.path.join(mapa, klasa)
        if not os.path.isdir(dir_klase):
            continue
        for f in os.listdir(dir_klase):
            if f.lower().endswith(PODRZANI_FORMATI):
                uzorci.append((os.path.join(dir_klase, f), oznaka))
    return uzorci


def _skupi_rsna():
    """RSNA: (putanja, Target) — sam nađe train-mapu i CSV s 'Target' u Data/.
    Dedup po patientId (1 = upala/opacity)."""
    import csv, glob
    podmape = [d for d in glob.glob(os.path.join(DIR_PODACI, "*")) if os.path.isdir(d)]
    slike_dir = next((d for d in podmape if "train" in os.path.basename(d).lower()), None) \
                or (podmape[0] if podmape else DIR_PODACI)
    po_stemu = {os.path.splitext(f)[0]: os.path.join(slike_dir, f)
                for f in os.listdir(slike_dir) if f.lower().endswith(PODRZANI_FORMATI)}
    csv_put = None
    for c in glob.glob(os.path.join(DIR_PODACI, "*.csv")):
        with open(c, newline="") as fh:
            if "Target" in fh.readline():
                csv_put = c; break
    if not csv_put:
        raise RuntimeError("RSNA: ne nalazim CSV s 'Target' u Data/.")
    labela = {}
    with open(csv_put, newline="") as fh:
        for row in csv.DictReader(fh):
            pid = row["patientId"]
            labela[pid] = max(labela.get(pid, 0), int(row.get("Target", 0)))
    return [(po_stemu[pid], t) for pid, t in labela.items() if pid in po_stemu]


def _pacijent_kljuc(putanja):
    """Grupiranje po pacijentu da isti pacijent ne bude i u train i u val.
    Pneumonija: imena tipa 'person78_bacteria_381' -> grupa 'person78'.
    Normalne (IM-...): nema pouzdanog ID-a -> svaka slika svoja grupa."""
    import re
    ime = os.path.basename(putanja)
    m = re.search(r"person\d+", ime, re.IGNORECASE)
    return m.group(0).lower() if m else os.path.splitext(ime)[0]


def napravi_dataloadere(img_size=None, batch=None, verbose=True):
    from torch.utils.data import WeightedRandomSampler
    img_size = img_size or IMG_SIZE
    batch = batch or BATCH

    # brži loaderi: više radnika + pin_memory + persistent_workers
    _kw = dict(num_workers=NUM_WORKERS, pin_memory=(DEVICE == "cuda"))
    if NUM_WORKERS > 0:
        _kw.update(prefetch_factor=4)

    if IZVOR_PODATAKA == "imagefolder":
        svi = _skupi_rsna()
        if not svi:
            raise RuntimeError("RSNA: nema uzoraka — provjeri Data/ (slike + CSV s Target).")
        from sklearn.model_selection import GroupShuffleSplit
        gr = [_pacijent_kljuc(p) for p, _ in svi]
        oz = [o for _, o in svi]; ix = list(range(len(svi)))
        gss1 = GroupShuffleSplit(n_splits=1, test_size=TEST_UDIO, random_state=SJEME)
        ost_idx, te_idx = next(gss1.split(ix, oz, groups=gr))
        test_uzorci = [svi[i] for i in te_idx]
        spojeno = [svi[i] for i in ost_idx]
    else:
        korijen = _rijesi_korijen(DIR_PODACI)
        spojeno = _skupi(os.path.join(korijen, "train")) + _skupi(os.path.join(korijen, "val"))
        test_uzorci = _skupi(os.path.join(korijen, "test"))
    if not spojeno:
        raise RuntimeError("Train skup je prazan — provjeri putanju do dataseta.")

    oznake = [o for _, o in spojeno]
    indeksi = list(range(len(spojeno)))
    grupe = [_pacijent_kljuc(p) for p, _ in spojeno]

    # Patient-grouped split (bez curenja pacijenta train<->val).
    try:
        from sklearn.model_selection import StratifiedGroupKFold
        skf = StratifiedGroupKFold(n_splits=max(2, round(1 / VAL_UDIO)),
                                   shuffle=True, random_state=SJEME)
        tr_idx, va_idx = next(skf.split(indeksi, oznake, groups=grupe))
    except Exception:
        from sklearn.model_selection import GroupShuffleSplit
        gss = GroupShuffleSplit(n_splits=1, test_size=VAL_UDIO, random_state=SJEME)
        tr_idx, va_idx = next(gss.split(indeksi, oznake, groups=grupe))

    train_uzorci = [spojeno[i] for i in tr_idx]
    val_uzorci = [spojeno[i] for i in va_idx]

    train_ds = XrayDataset(train_uzorci, napravi_transforme(img_size, augmentacija=True))
    val_ds = XrayDataset(val_uzorci, napravi_transforme(img_size, augmentacija=False))
    test_ds = XrayDataset(test_uzorci, napravi_transforme(img_size, augmentacija=False))

    # Mali train podskup (bez augmentacije) za mjerenje train metrika svaku epohu
    # -> da se train↔val rascjep (overfitting) vidi odmah.
    rng = np.random.default_rng(SJEME)
    n_eval = min(1000, len(train_uzorci))
    eval_idx = rng.choice(len(train_uzorci), size=n_eval, replace=False)
    train_eval_uzorci = [train_uzorci[i] for i in eval_idx]
    train_eval_ds = XrayDataset(train_eval_uzorci, napravi_transforme(img_size, augmentacija=False))

    train_oznake = [o for _, o in train_uzorci]
    brojevi = np.bincount(train_oznake, minlength=len(KLASE))

    # Balans klasa
    if BALANS_KLASA == "sampler":
        w_po_klasi = 1.0 / np.maximum(brojevi, 1)
        w_uzorka = [float(w_po_klasi[o]) for o in train_oznake]
        sampler = WeightedRandomSampler(torch.as_tensor(w_uzorka, dtype=torch.float64),
                                        num_samples=len(w_uzorka), replacement=True)
        train_loader = DataLoader(train_ds, batch_size=batch, sampler=sampler, **_kw)
        tezine_klasa = None                      # sampler vec balansira -> loss bez tezina
    else:
        train_loader = DataLoader(train_ds, batch_size=batch, shuffle=True, **_kw)
        if BALANS_KLASA == "loss":
            tezine = len(train_uzorci) / (len(KLASE) * np.maximum(brojevi, 1))
            tezine_klasa = torch.tensor(tezine, dtype=torch.float32)
        else:
            tezine_klasa = None

    val_loader = DataLoader(val_ds, batch_size=batch, shuffle=False, **_kw)
    test_loader = DataLoader(test_ds, batch_size=batch, shuffle=False, **_kw)
    train_eval_loader = DataLoader(train_eval_ds, batch_size=batch, shuffle=False, **_kw)

    if verbose:
        print(f"Train: {len(train_ds)} | Val: {len(val_ds)} | Test: {len(test_ds)} "
              f"| train-eval podskup: {len(train_eval_ds)}")
        print(f"Balans klasa: {BALANS_KLASA} · raspodjela train [NORMAL, PNEU]: {brojevi.tolist()}")
        # provjera curenja: presjek pacijenata train/val mora biti prazan
        gtr = {grupe[i] for i in tr_idx}
        gva = {grupe[i] for i in va_idx}
        print(f"Zajednickih pacijenata train∩val: {len(gtr & gva)} (treba 0)")
    return train_loader, val_loader, test_loader, tezine_klasa, train_eval_loader

# ================================================================
# MODELI
# ================================================================
class SEBlok(nn.Module):
    """Squeeze-and-Excitation: nauči koliko 'pojačati' svaki kanal (channel attention).
    Jeftino, smanjuje oslanjanje na nevažne mape -> blago bolja generalizacija."""
    def __init__(self, kanali, redukcija=16):
        super().__init__()
        sred = max(1, kanali // redukcija)
        self.avg = nn.AdaptiveAvgPool2d(1)
        self.fc = nn.Sequential(
            nn.Linear(kanali, sred), nn.ReLU(inplace=True),
            nn.Linear(sred, kanali), nn.Sigmoid(),
        )

    def forward(self, x):
        b, c, _, _ = x.shape
        s = self.avg(x).view(b, c)
        s = self.fc(s).view(b, c, 1, 1)
        return x * s


class CNNOdNule(nn.Module):
    def __init__(self, broj_klasa=2):
        super().__init__()

        def blok(ulaz, izlaz):
            return nn.Sequential(
                nn.Conv2d(ulaz, izlaz, 3, padding=1),
                nn.BatchNorm2d(izlaz),
                nn.ReLU(inplace=True),
                nn.Conv2d(izlaz, izlaz, 3, padding=1),
                nn.BatchNorm2d(izlaz),
                nn.ReLU(inplace=True),
                SEBlok(izlaz),            # channel attention
                nn.Dropout2d(0.1),        # spatial dropout (regularizacija na mapama)
                nn.MaxPool2d(2),
            )

        self.features = nn.Sequential(blok(3, 32), blok(32, 64), blok(64, 128), blok(128, 256))
        self.pool = nn.AdaptiveAvgPool2d(1)
        self.classifier = nn.Sequential(
            nn.Flatten(), nn.Dropout(0.4), nn.Linear(256, 128), nn.ReLU(inplace=True),
            nn.Dropout(0.3), nn.Linear(128, broj_klasa)
        )

    def forward(self, x):
        x = self.features(x)
        x = self.pool(x)
        return self.classifier(x)


def napravi_vgg16(broj_klasa=2, zamrzni_bazu=True):
    model = models.vgg16(weights=VGG16_Weights.IMAGENET1K_V1)
    if zamrzni_bazu:
        for p in model.features.parameters():
            p.requires_grad = False
    ulaz = model.classifier[6].in_features
    model.classifier[6] = nn.Linear(ulaz, broj_klasa)
    return model


def napravi_model(tip, broj_klasa=2, zamrzni_bazu=None):
    tip = tip.lower()
    if tip == "cnn":
        return CNNOdNule(broj_klasa)
    if tip == "vgg16":
        if zamrzni_bazu is None:
            zamrzni_bazu = ZAMRZNI_VGG_BAZU
        return napravi_vgg16(broj_klasa, zamrzni_bazu)
    raise ValueError(f"Nepoznat tip modela: {tip!r}.")


def odmrzni_vgg(model, n_blokova=1):
    """Odmrzni zadnjih n conv blokova VGG baze (blok = sloj do MaxPool2d).
    Vraća broj odmrznutih modula. Koristi se za fazu 2 fine-tuninga."""
    feats = getattr(model, "features", None)
    if feats is None:
        return 0
    pool_idx = [i for i, m in enumerate(feats) if isinstance(m, nn.MaxPool2d)]
    if not pool_idx:
        return 0
    granica = -1 if n_blokova >= len(pool_idx) else pool_idx[-(n_blokova + 1)]
    n = 0
    for i, mod in enumerate(feats):
        if i > granica:
            for p in mod.parameters():
                p.requires_grad = True
                n += 1
    return n


def spremi_model(model, putanja, meta=None):
    os.makedirs(os.path.dirname(putanja) or ".", exist_ok=True)
    torch.save({"state_dict": model.state_dict(), "meta": meta or {}}, putanja)


def ucitaj_model(tip, putanja=None, device=None):
    device = device or DEVICE
    putanja = putanja or PUTANJA_MODEL[tip.lower()]
    model = napravi_model(tip)
    try:
        paket = torch.load(putanja, map_location=device, weights_only=False)
    except TypeError:
        paket = torch.load(putanja, map_location=device)
    state = paket["state_dict"] if isinstance(paket, dict) and "state_dict" in paket else paket
    meta = paket.get("meta", {}) if isinstance(paket, dict) else {}
    model.load_state_dict(state)
    model.to(device).eval()
    return model, meta

# ================================================================
# (Metrike, evaluacija i usporedba modela preseljeni su u evaluacija.py)
# (Treniranje je preseljeno u treniranje.py)
# Donji dio drži zajedničku logiku: podatke, modele, Grad-CAM,
# inferenciju i izvještaje — koje koriste i GUI i CLI.
# ================================================================

# ================================================================
# GRAD-CAM
# ================================================================
def _colormap(naziv):
    try:
        from matplotlib import colormaps
        return colormaps[naziv]
    except Exception:
        import matplotlib.cm as cm
        return cm.get_cmap(naziv)


def zadnji_konv_sloj(model):
    zadnji = None
    for modul in model.modules():
        if isinstance(modul, nn.Conv2d):
            zadnji = modul
    if zadnji is None:
        raise RuntimeError("Model nema nijedan Conv2d sloj.")
    return zadnji


class GradCAM:
    def __init__(self, model, ciljni_sloj=None):
        self.model = model
        self.ciljni_sloj = ciljni_sloj or zadnji_konv_sloj(model)
        self.aktivacije = None
        self.gradijenti = None
        self._h = self.ciljni_sloj.register_forward_hook(self._fwd_hook)

    def _fwd_hook(self, modul, ulaz, izlaz):
        self.aktivacije = izlaz
        izlaz.register_hook(self._grad_hook)

    def _grad_hook(self, grad):
        self.gradijenti = grad

    def generiraj(self, ulaz_tensor, klasa=None):
        self.model.eval()
        ulaz = ulaz_tensor.clone().detach().requires_grad_(True)
        with torch.enable_grad():
            izlaz = self.model(ulaz)
            prob = F.softmax(izlaz, dim=1)
            predvidjena = int(izlaz.argmax(dim=1).item())
            cilj = predvidjena if klasa is None else klasa
            self.model.zero_grad(set_to_none=True)
            izlaz[0, cilj].backward()
        tezine = self.gradijenti.mean(dim=(2, 3), keepdim=True)
        cam = (tezine * self.aktivacije).sum(dim=1, keepdim=True)
        cam = F.relu(cam)
        cam = F.interpolate(cam, size=ulaz.shape[-2:], mode="bilinear", align_corners=False)
        cam = cam[0, 0]
        cam = cam - cam.min()
        if cam.max() > 0:
            cam = cam / cam.max()
        return cam.detach().cpu().numpy(), predvidjena, prob.detach().cpu().numpy()[0]

    def ukloni(self):
        self._h.remove()


def primijeni_colormap(heatmap, mapa="jet"):
    boje = _colormap(mapa)(heatmap)[:, :, :3]
    return (boje * 255).astype(np.uint8)


def preklopi(original_pil, heatmap, alpha=0.45, mapa="jet"):
    baza = original_pil.convert("RGB").resize((heatmap.shape[1], heatmap.shape[0]))
    heat_rgb = Image.fromarray(primijeni_colormap(heatmap, mapa=mapa)).convert("RGB")
    return Image.blend(baza, heat_rgb, alpha=alpha)

# ================================================================
# INFERENCIJA I IZVJEŠTAJI
# ================================================================
_model = None
_tip = None
_prag = PRAG_ODLUKE
_transform = transform_inferencija()


def ucitaj(tip, putanja=None, device=None):
    global _model, _tip, _prag
    device = device or DEVICE
    model, meta = ucitaj_model(tip, putanja, device)
    _model = model
    _tip = tip
    _prag = meta.get("prag", PRAG_ODLUKE)
    return _tip, _prag


def model_ucitan():
    return _model is not None


def trenutni_tip():
    return _tip


def trenutni_prag():
    return _prag


def prikupi_slike(putanje):
    izlaz = []
    for p in putanje:
        if os.path.isdir(p):
            for f in sorted(os.listdir(p)):
                if f.lower().endswith(PODRZANI_FORMATI):
                    izlaz.append(os.path.join(p, f))
        elif os.path.isfile(p) and p.lower().endswith(PODRZANI_FORMATI):
            izlaz.append(p)
    vidjeno, jedinstveno = set(), []
    for p in izlaz:
        if p not in vidjeno:
            vidjeno.add(p)
            jedinstveno.append(p)
    return jedinstveno


def obradi_snimku(putanja, prag=None, gradcam_klasa=None, device=None):
    if _model is None:
        raise RuntimeError("Model nije učitan — prvo pozovi ucitaj(tip).")
    device = device or DEVICE
    prag = _prag if prag is None else prag
    original = Image.open(putanja).convert("L")
    import segmentacija
    original = segmentacija.mozda_izrezi(original, putanja)
    tensor = _transform(original).unsqueeze(0).to(device)

    if gradcam_klasa is None:
        klasa_param = POZITIVNA_KLASA if GRADCAM_KLASA == "pneumonija" else None
    elif gradcam_klasa == "pneumonija":
        klasa_param = POZITIVNA_KLASA
    else:
        klasa_param = None

    gc = GradCAM(_model)
    try:
        heatmap, predvidjena, prob = gc.generiraj(tensor, klasa=klasa_param)
    finally:
        gc.ukloni()

    p_pneumonija = float(prob[POZITIVNA_KLASA])
    p_normal = float(prob[1 - POZITIVNA_KLASA])
    oznaka = POZITIVNA_KLASA if p_pneumonija >= prag else (1 - POZITIVNA_KLASA)
    # baza za prikaz/overlay = ISTA geometrija koju model vidi -> poravnat heatmap
    baza = geometrija_prikaza(original)
    overlay = preklopi(baza, heatmap, alpha=GRADCAM_ALPHA)
    prikaz = KLASE_PRIKAZ

    return {"naziv": os.path.basename(putanja), "putanja": putanja, "predikcija": prikaz[oznaka],
            "oznaka": oznaka, "p_pneumonija": p_pneumonija, "p_normal": p_normal,
            "prag": prag, "original": baza, "overlay": overlay}


def analiziraj(putanje, na_napredak=None, gradcam_klasa=None):
    slike = prikupi_slike(putanje)
    rez = []
    for i, p in enumerate(slike, start=1):
        rez.append(obradi_snimku(p, gradcam_klasa=gradcam_klasa))
        if na_napredak:
            na_napredak(i, len(slike), os.path.basename(p))
    return rez


def spremi_izvjestaj_txt(rezultati, putanja, timestamp=None):
    timestamp = timestamp or datetime.datetime.now().strftime("%d.%m.%Y. u %H:%M:%S")
    with open(putanja, "w", encoding="utf-8") as f:
        f.write("IZVJEŠTAJ — Detekcija upale pluća iz rendgenskih snimki\n")
        f.write(f"Vrijeme analize: {timestamp}\n")
        f.write(f"Model: {(_tip or '?').upper()}   Prag odluke: {_prag:.3f}\n")
        f.write("=" * 60 + "\n\n")
        n_pneu = sum(1 for r in rezultati if r["oznaka"] == POZITIVNA_KLASA)
        f.write(f"Ukupno snimki: {len(rezultati)}   UPALA PLUĆA: {n_pneu}   NORMALNO: {len(rezultati)-n_pneu}\n\n")
        for i, r in enumerate(rezultati, start=1):
            f.write(f"{i}. {r['naziv']}\n")
            f.write(f"    Predikcija   : {r['predikcija']}\n")
            f.write(f"    P(upala)     : {r['p_pneumonija']*100:.2f} %\n")
            f.write(f"    P(normalno)  : {r['p_normal']*100:.2f} %\n")
            f.write(f"    Prag odluke  : {r['prag']:.3f}\n\n")


def spremi_izvjestaj_xlsx(rezultati, putanja, timestamp=None, sa_slikama=True):
    """Excel nalaz oblikovan da prati GUI: naslov, sažetak, obojeni 'Nalaz'
    (UPALA crveno / NORMALNO zeleno), metrika ćelije i ugrađene slike."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    # GUI paleta
    C_TEXT = "1F2630"; C_MUTE = "6B7280"; C_PRIMARY = "4B5563"; C_BORDER = "E3E6EA"
    C_CARDALT = "F8F9FB"; C_OK = "0E9F6E"; C_OKBG = "E6F6EE"
    C_DANGER = "E11D48"; C_DANGERBG = "FFE4EA"
    F = "Segoe UI"

    timestamp = timestamp or datetime.datetime.now().strftime("%d.%m.%Y. u %H:%M:%S")
    wb = Workbook(); ws = wb.active; ws.title = "Nalaz"
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sirine = {"A": 5, "B": 26, "C": 23, "D": 23, "E": 22, "F": 12, "G": 13, "H": 9, "I": 11}
    for col, w in sirine.items():
        ws.column_dimensions[col].width = w

    n = len(rezultati)
    n_up = sum(1 for r in rezultati if r["oznaka"] == POZITIVNA_KLASA)

    # 1) naslov
    ws.merge_cells("A1:I1")
    ws["A1"] = "Sustav za detekciju upale pluća iz rendgenskih snimki"
    ws["A1"].font = Font(name=F, bold=True, size=15, color=C_TEXT)
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    for col in range(1, 10):
        ws.cell(1, col).border = Border(bottom=Side(style="thin", color=C_BORDER))

    # 2) sažetak (obojeno)
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Analiza završena: upala pluća detektirana na {n_up} od ukupno {n} snimki"
    ws["A2"].font = Font(name=F, bold=True, size=12, color=(C_DANGER if n_up else C_OK))
    ws["A2"].fill = PatternFill("solid", fgColor=(C_DANGERBG if n_up else C_OKBG))
    ws["A2"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 24

    # 3) info redak
    ws.merge_cells("A3:I3")
    ws["A3"] = (f"Vrijeme: {timestamp}    ·    Model: {(_tip or '?').upper()}"
                f"    ·    Prag odluke: {_prag:.3f}")
    ws["A3"].font = Font(name=F, size=10, color=C_MUTE)
    ws["A3"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[3].height = 18

    # 4) zaglavlje tablice
    head = 4
    zagl = ["#", "Snimka", "Originalna snimka", "Grad-CAM toplinska karta", "Nalaz",
            "P(upala)", "P(normalno)", "Prag", "Model"]
    for ci, h in enumerate(zagl, 1):
        c = ws.cell(head, ci, h)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=C_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[head].height = 30
    ws.freeze_panes = f"A{head + 1}"

    tmp_images = []
    r0 = head + 1
    for i, r in enumerate(rezultati):
        rr = r0 + i
        upala = r["oznaka"] == POZITIVNA_KLASA
        akc = C_DANGER if upala else C_OK
        akcbg = C_DANGERBG if upala else C_OKBG
        p = r["p_pneumonija"] if upala else r["p_normal"]
        natpis = "⚠  UPALA PLUĆA" if upala else "✅  NORMALNO"

        ws.cell(rr, 1, i + 1).alignment = Alignment(horizontal="center", vertical="center")
        b = ws.cell(rr, 2, r["naziv"])
        b.font = Font(name=F, size=10, color=C_TEXT, bold=True)
        b.alignment = Alignment(vertical="center", wrap_text=True, indent=1)

        cn = ws.cell(rr, 5, f"{natpis}\n{p*100:.1f}%")
        cn.font = Font(name=F, bold=True, size=11, color=akc)
        cn.fill = PatternFill("solid", fgColor=akcbg)
        cn.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col, val, bold in [(6, f"{r['p_pneumonija']*100:.1f}%", upala),
                               (7, f"{r['p_normal']*100:.1f}%", not upala),
                               (8, f"{r['prag']:.3f}", False),
                               (9, (_tip or '?').upper(), False)]:
            cc = ws.cell(rr, col, val)
            cc.font = Font(name=F, bold=bool(bold), size=10, color=C_TEXT)
            cc.fill = PatternFill("solid", fgColor=C_CARDALT)
            cc.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 10):
            ws.cell(rr, col).border = border
        ws.row_dimensions[rr].height = 120 if sa_slikama else 22

        if sa_slikama:
            for col, key in [(3, "original"), (4, "overlay")]:
                bio = io.BytesIO()
                im = r[key].copy()
                im.thumbnail((155, 155))
                im.save(bio, format="PNG")
                bio.seek(0)
                ws.add_image(XLImage(bio), f"{get_column_letter(col)}{rr}")
                tmp_images.append(bio)

    os.makedirs(os.path.dirname(putanja) or ".", exist_ok=True)
    wb.save(putanja)


def spremi_heatmape(rezultati, mapa):
    os.makedirs(mapa, exist_ok=True)
    for i, r in enumerate(rezultati, start=1):
        baza = os.path.splitext(r["naziv"])[0]
        safe = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in baza)
        r["overlay"].save(os.path.join(mapa, f"{i:02d}_{safe}_gradcam.png"))


# ================================================================
# POKRETAČI
# ================================================================
def pokreni_gui():
    from gui import App
    App().mainloop()


def pokreni_cli(putanje, tip="vgg16", model_putanja=None, spremi="txt", izlaz=None, gradcam_klasa=None):
    import datetime
    import treniranje as ui

    izlaz = izlaz or DIR_REZULTATI
    os.makedirs(izlaz, exist_ok=True)
    ui.hint_rich()

    slike = prikupi_slike(putanje)
    if not slike:
        ui.greska("Nije pronađena nijedna podržana snimka.")
        ui.info(f"Podržani formati: {', '.join(PODRZANI_FORMATI)}")
        return

    ui.naslov("Sustav za detekciju upale pluća", "terminalni način rada · iz rendgenskih snimki")
    putanja_modela = model_putanja or PUTANJA_MODEL[tip]
    if not os.path.exists(putanja_modela):
        ui.greska(f"Model ne postoji: {putanja_modela}")
        ui.info(f"Prvo istreniraj: python main.py train {tip}")
        return

    ui.info(f"Učitavam model: {tip.upper()} ({putanja_modela})")
    _, prag = ucitaj(tip, putanja_modela)
    ui.uspjeh(f"Model spreman · uređaj: {DEVICE} · prag odluke: {prag:.3f}")

    rezultati = []
    with ui.traka("Analiza", len(slike)) as bar:
        for p in slike:
            r = obradi_snimku(p, gradcam_klasa=gradcam_klasa)
            rezultati.append(r)
            bar.korak(info=os.path.basename(p))

    redovi = []
    for r in rezultati:
        upala = r["oznaka"] == POZITIVNA_KLASA
        redovi.append([
            r["naziv"],
            ("⚠ " if upala else "✓ ") + r["predikcija"],
            f"{r['p_pneumonija'] * 100:.1f}%",
            f"{r['p_normal'] * 100:.1f}%",
        ])
    ui.tablica("Rezultati dijagnostike", ["Snimka", "Predikcija", "P(upala)", "P(normalno)"], redovi)

    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if spremi:
        ext = "xlsx" if spremi == "xlsx" else "txt"
        out = os.path.join(izlaz, f"nalaz_{ts}.{ext}")
        if ext == "xlsx":
            spremi_izvjestaj_xlsx(rezultati, out)
        else:
            spremi_izvjestaj_txt(rezultati, out)
        mapa_png = os.path.join(izlaz, f"gradcam_{ts}")
        spremi_heatmape(rezultati, mapa_png)
        ui.uspjeh(f"Spremljeno: {out}")
        ui.info(f"Grad-CAM slike: {mapa_png}")


def main():
    parser = argparse.ArgumentParser(description="Detekcija upale pluća iz rendgenskih snimki")
    sub = parser.add_subparsers(dest="cmd")

    sub.add_parser("gui", help="pokreni grafičko sučelje")

    p_cli = sub.add_parser("cli", help="terminalna analiza snimki")
    p_cli.add_argument("putanje", nargs="+", help="slike ili mape")
    p_cli.add_argument("--model", choices=["cnn", "vgg16"], default="vgg16")
    p_cli.add_argument("--model-putanja", default=None)
    p_cli.add_argument("--spremi", choices=["txt", "xlsx", "ne"], default="txt")
    p_cli.add_argument("--izlaz", default=None)
    p_cli.add_argument("--gradcam-klasa", choices=["predvidjena", "pneumonija"], default=None)

    p_train = sub.add_parser("train", help="treniranje modela")
    p_train.add_argument("tip", nargs="?", choices=["cnn", "vgg16", "oba"], default="oba")

    sub.add_parser("eval", help="usporedba spremljenih modela")

    args = parser.parse_args()

    if args.cmd in (None, "gui"):
        pokreni_gui()
    elif args.cmd == "cli":
        pokreni_cli(args.putanje, args.model, args.model_putanja,
                   None if args.spremi == "ne" else args.spremi,
                   args.izlaz, args.gradcam_klasa)
    elif args.cmd == "train":
        import treniranje
        treniranje.treniraj_iz_cli(args.tip)
    elif args.cmd == "eval":
        import evaluacija
        evaluacija.usporedi()


if __name__ == "__main__":
    main()
